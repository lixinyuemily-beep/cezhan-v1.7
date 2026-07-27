"""
展品固定模板解析服务
"""
import asyncio
import csv
import io
import shutil
import uuid
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError

from ..config import settings
from .storage_service import storage_service


EXPECTED_HEADERS = ["展品名称", "时间", "地点", "材质", "介绍", "图片", "其他"]
REQUIRED_HEADERS = ["展品名称", "时间", "地点", "材质", "介绍", "图片"]
IMAGE_COLUMN_INDEX = EXPECTED_HEADERS.index("图片")
MAX_UPLOAD_SIZE_BYTES = settings.exhibit_import_max_upload_size_mb * 1024 * 1024
UPLOAD_CHUNK_SIZE_BYTES = 1024 * 1024
PREVIEW_MAX_EDGE = 1280
THUMBNAIL_MAX_EDGE = 320


class ExhibitImportService:
    """异步解析固定模板的展品表格，并尽可能提取图片内容。"""

    _tasks: Dict[str, Dict[str, Any]] = {}
    _task_lock = asyncio.Lock()
    _parse_semaphore = asyncio.Semaphore(max(1, settings.exhibit_import_task_concurrency))

    @staticmethod
    def _temp_root() -> Path:
        return Path(settings.exhibit_import_temp_dir)

    @classmethod
    def _ensure_storage_dirs(cls) -> None:
        cls._temp_root().mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _normalize_header(value) -> str:
        return str(value or "").strip()

    @classmethod
    def _validate_headers(cls, headers: List[str]) -> List[str]:
        normalized = [cls._normalize_header(header) for header in headers]
        while normalized and normalized[-1] == "":
            normalized.pop()

        if normalized != EXPECTED_HEADERS:
            expected = "、".join(EXPECTED_HEADERS)
            current = "、".join(normalized) if normalized else "空表头"
            raise HTTPException(
                status_code=400,
                detail=f"上传失败：模板表头必须严格为「{expected}」，当前为「{current}」。",
            )
        return normalized

    @staticmethod
    def _stringify(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @classmethod
    def _is_effectively_blank_row(cls, row_dict: Dict[str, str]) -> bool:
        return not any(cls._stringify(row_dict.get(header)) for header in EXPECTED_HEADERS)

    @classmethod
    def _get_missing_fields(cls, row_dict: Dict[str, str]) -> List[str]:
        return [header for header in REQUIRED_HEADERS if not cls._stringify(row_dict.get(header))]

    @classmethod
    def _build_exhibit_record(cls, row_dict: Dict[str, str]) -> Tuple[Dict[str, str], List[str]]:
        name = cls._stringify(row_dict.get("展品名称"))
        missing_fields = cls._get_missing_fields(row_dict)

        time_value = cls._stringify(row_dict.get("时间"))
        place = cls._stringify(row_dict.get("地点"))
        material = cls._stringify(row_dict.get("材质"))
        introduction = cls._stringify(row_dict.get("介绍"))
        image_url = cls._stringify(row_dict.get("图片"))
        thumbnail_url = cls._stringify(row_dict.get("缩略图")) or image_url
        storage_bucket = cls._stringify(row_dict.get("存储桶"))
        storage_path = cls._stringify(row_dict.get("存储路径"))
        thumbnail_storage_path = cls._stringify(row_dict.get("缩略图存储路径"))
        other = cls._stringify(row_dict.get("其他"))

        return {
            "name": name,
            "time": time_value,
            "place": place,
            "material": material,
            "introduction": introduction,
            "image_url": image_url,
            "thumbnail_url": thumbnail_url,
            "storage_bucket": storage_bucket,
            "storage_path": storage_path,
            "thumbnail_storage_path": thumbnail_storage_path,
            "other": other,
            # 向后兼容旧前端/旧项目数据读取逻辑
            "era": time_value,
            "description": introduction,
            "size": other,
        }, missing_fields

    @classmethod
    async def _update_task(cls, task_id: str, **updates: Any) -> None:
        async with cls._task_lock:
            task = cls._tasks.get(task_id)
            if not task:
                return
            task.update(updates)

    @classmethod
    async def get_parse_task(cls, task_id: str, user_id: str) -> Dict[str, Any]:
        async with cls._task_lock:
            task = cls._tasks.get(task_id)
            if not task:
                raise HTTPException(status_code=404, detail="解析任务不存在或已过期。")
            if task.get("user_id") != user_id:
                raise HTTPException(status_code=403, detail="不能访问其他用户的解析任务。")
            return dict(task)

    @classmethod
    async def create_parse_task(cls, file: UploadFile, user_id: str) -> Dict[str, Any]:
        cls._ensure_storage_dirs()

        filename = str(file.filename or "").strip()
        lower_name = filename.lower()
        if not filename:
            raise HTTPException(status_code=400, detail="上传失败：未检测到文件名。")
        if not (lower_name.endswith(".xlsx") or lower_name.endswith(".csv")):
            raise HTTPException(status_code=400, detail="上传失败：目前仅支持固定模板的 .xlsx 或 .csv 文件。")

        task_id = uuid.uuid4().hex
        task_dir = cls._temp_root() / task_id
        task_dir.mkdir(parents=True, exist_ok=True)
        source_path = task_dir / f"source{Path(filename).suffix.lower()}"

        total_size = 0
        try:
            with source_path.open("wb") as target:
                while True:
                    chunk = await file.read(UPLOAD_CHUNK_SIZE_BYTES)
                    if not chunk:
                        break
                    total_size += len(chunk)
                    if total_size > MAX_UPLOAD_SIZE_BYTES:
                        raise HTTPException(
                            status_code=400,
                            detail=f"上传失败：文件超过 {settings.exhibit_import_max_upload_size_mb}MB，请压缩图片或拆分后再上传。",
                        )
                    target.write(chunk)
        finally:
            await file.close()

        if total_size == 0:
            raise HTTPException(status_code=400, detail="上传失败：文件为空。")

        task = {
            "task_id": task_id,
            "user_id": user_id,
            "status": "queued",
            "stage": "queued",
            "progress": 0,
            "message": "文件上传完成，等待开始解析。",
            "file_name": filename,
            "meta": {
                "file_size": total_size,
                "embedded_image_count": 0,
            },
            "result": None,
            "error": None,
        }
        async with cls._task_lock:
            cls._tasks[task_id] = task

        asyncio.create_task(cls._process_task(task_id, source_path, filename, user_id))
        return task

    @classmethod
    async def _process_task(cls, task_id: str, source_path: Path, filename: str, user_id: str) -> None:
        event_loop = asyncio.get_running_loop()
        uploaded_storage_paths: List[str] = []

        def report(stage: str, progress: int, message: str) -> None:
            event_loop.call_soon_threadsafe(
                asyncio.create_task,
                cls._update_task(
                    task_id,
                    status="processing",
                    stage=stage,
                    progress=max(0, min(progress, 99)),
                    message=message,
                ),
            )

        try:
            await cls._update_task(
                task_id,
                status="processing",
                stage="queued",
                progress=2,
                message="解析任务已创建，等待处理。",
            )
            async with cls._parse_semaphore:
                report("validating", 8, "正在校验文件格式和模板表头。")
                result = await asyncio.to_thread(
                    cls._parse_saved_file,
                    source_path,
                    filename,
                    task_id,
                    user_id,
                    uploaded_storage_paths,
                    report,
                )
                await cls._update_task(
                    task_id,
                    status="success",
                    stage="success",
                    progress=100,
                    message="解析完成，可查看展品清单。",
                    meta=result.get("meta"),
                    result=result,
                    error=None,
                )
        except HTTPException as exc:
            await asyncio.to_thread(storage_service.remove_files, uploaded_storage_paths)
            await cls._update_task(
                task_id,
                status="failed",
                stage="failed",
                progress=100,
                message="解析失败。",
                error=exc.detail,
                result=None,
            )
        except Exception as exc:
            await asyncio.to_thread(storage_service.remove_files, uploaded_storage_paths)
            await cls._update_task(
                task_id,
                status="failed",
                stage="failed",
                progress=100,
                message="解析失败。",
                error=f"上传失败：解析过程中发生异常，{exc}",
                result=None,
            )
        finally:
            shutil.rmtree(source_path.parent, ignore_errors=True)

    @classmethod
    def _encode_image_bytes(cls, image: Image.Image, max_edge: int) -> Tuple[bytes, str, str]:
        working = image.copy()
        working.thumbnail((max_edge, max_edge))
        output = io.BytesIO()

        has_alpha = "A" in working.getbands()
        if has_alpha:
            if working.mode not in ("RGBA", "LA"):
                working = working.convert("RGBA")
            working.save(output, format="PNG", optimize=True)
            return output.getvalue(), "png", "image/png"

        if working.mode != "RGB":
            working = working.convert("RGB")
        working.save(output, format="JPEG", quality=82, optimize=True)
        return output.getvalue(), "jpg", "image/jpeg"

    @classmethod
    def _optimize_embedded_image(
        cls,
        image_bytes: bytes,
        user_id: str,
        task_id: str,
        uploaded_storage_paths: List[str],
    ) -> Tuple[str, str, str, str]:
        try:
            image = Image.open(io.BytesIO(image_bytes))
            image.load()
        except (UnidentifiedImageError, OSError) as exc:
            raise HTTPException(status_code=400, detail=f"上传失败：存在无法识别的嵌入图片，{exc}") from exc

        image_id = uuid.uuid4().hex
        preview_content, preview_ext, preview_content_type = cls._encode_image_bytes(image, PREVIEW_MAX_EDGE)
        thumbnail_content, thumbnail_ext, thumbnail_content_type = cls._encode_image_bytes(image, THUMBNAIL_MAX_EDGE)

        preview_name = f"{image_id}.{preview_ext}"
        thumbnail_name = f"{image_id}_thumb.{thumbnail_ext}"
        preview_path = storage_service.build_import_asset_path(user_id, task_id, preview_name)
        thumbnail_path = storage_service.build_import_asset_path(user_id, task_id, thumbnail_name)

        preview_url = storage_service.upload_bytes(preview_path, preview_content, preview_content_type)
        thumbnail_url = storage_service.upload_bytes(thumbnail_path, thumbnail_content, thumbnail_content_type)
        uploaded_storage_paths.extend([preview_path, thumbnail_path])
        return preview_url, thumbnail_url, preview_path, thumbnail_path

    @classmethod
    def _parse_csv(cls, file_path: Path) -> Tuple[List[str], List[Dict[str, str]], Dict[str, Any]]:
        content = file_path.read_bytes()
        decoded = None
        for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded is None:
            raise HTTPException(status_code=400, detail="上传失败：CSV 文件编码无法识别，请使用 UTF-8 或 GBK。")

        reader = csv.reader(io.StringIO(decoded))
        rows = list(reader)
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="上传失败：文件内容为空，至少需要表头和一行数据。")

        headers = cls._validate_headers(rows[0])
        exhibits = []
        incomplete_rows = []
        skipped_blank_rows = 0
        for index, row in enumerate(rows[1:], start=2):
            normalized_row = list(row[: len(headers)])
            if len(normalized_row) < len(headers):
                normalized_row += [""] * (len(headers) - len(normalized_row))

            row_dict = dict(zip(headers, normalized_row))
            if cls._is_effectively_blank_row(row_dict):
                skipped_blank_rows += 1
                continue
            exhibit, missing_fields = cls._build_exhibit_record(row_dict)
            exhibits.append(exhibit)
            if missing_fields:
                incomplete_rows.append({
                    "row_number": index,
                    "name": exhibit["name"],
                    "missing_fields": missing_fields,
                })

        if not exhibits:
            raise HTTPException(status_code=400, detail="上传失败：未解析到有效展品数据。")

        return headers, exhibits, {
            "embedded_image_count": 0,
            "skipped_blank_rows": skipped_blank_rows,
            "incomplete_rows": incomplete_rows,
            "incomplete_row_count": len(incomplete_rows),
        }

    @classmethod
    def _extract_sheet_images(
        cls,
        worksheet,
        user_id: str,
        task_id: str,
        uploaded_storage_paths: List[str],
        report: Callable[[str, int, str], None],
    ) -> Tuple[Dict[int, Dict[str, str]], int]:
        row_to_image: Dict[int, Dict[str, str]] = {}
        processed_count = 0
        images = getattr(worksheet, "_images", []) or []
        total_images = len(images)

        for index, image in enumerate(images, start=1):
            anchor = getattr(image, "anchor", None)
            marker = getattr(anchor, "_from", None)
            if marker is None:
                continue

            row_index = int(getattr(marker, "row", -1)) + 1
            col_index = int(getattr(marker, "col", -1))
            if row_index <= 0 or col_index != IMAGE_COLUMN_INDEX:
                continue

            try:
                image_bytes = image._data()
            except Exception:
                continue

            preview_url, thumbnail_url, storage_path, thumbnail_storage_path = cls._optimize_embedded_image(
                image_bytes,
                user_id,
                task_id,
                uploaded_storage_paths,
            )
            row_to_image[row_index] = {
                "image_url": preview_url,
                "thumbnail_url": thumbnail_url,
                "storage_bucket": settings.supabase_storage_bucket,
                "storage_path": storage_path,
                "thumbnail_storage_path": thumbnail_storage_path,
            }
            processed_count += 1

            if total_images:
                progress = 45 + int(index / total_images * 20)
                report("processing_images", progress, f"正在处理嵌入图片（{index}/{total_images}）。")

        return row_to_image, processed_count

    @classmethod
    def _parse_xlsx(
        cls,
        file_path: Path,
        user_id: str,
        task_id: str,
        uploaded_storage_paths: List[str],
        report: Callable[[str, int, str], None],
    ) -> Tuple[List[str], List[Dict[str, str]], Dict[str, int]]:
        try:
            workbook = load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"上传失败：无法读取 Excel 文件，{exc}") from exc

        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="上传失败：文件内容为空，至少需要表头和一行数据。")

        headers = cls._validate_headers(list(rows[0][: len(EXPECTED_HEADERS)]))
        report("reading_rows", 35, "模板校验通过，正在读取表格行数据。")
        image_by_row, embedded_image_count = cls._extract_sheet_images(
            worksheet,
            user_id,
            task_id,
            uploaded_storage_paths,
            report,
        )

        exhibits = []
        incomplete_rows = []
        skipped_blank_rows = 0
        total_rows = max(len(rows) - 1, 1)
        for index, row in enumerate(rows[1:], start=1):
            row_number = index + 1
            values = list(row[: len(headers)])
            if len(values) < len(headers):
                values += [""] * (len(headers) - len(values))

            row_dict = dict(zip(headers, values))
            if row_number in image_by_row:
                row_dict["图片"] = image_by_row[row_number]["image_url"]
                row_dict["缩略图"] = image_by_row[row_number]["thumbnail_url"]
                row_dict["存储桶"] = image_by_row[row_number]["storage_bucket"]
                row_dict["存储路径"] = image_by_row[row_number]["storage_path"]
                row_dict["缩略图存储路径"] = image_by_row[row_number]["thumbnail_storage_path"]
            if cls._is_effectively_blank_row(row_dict):
                skipped_blank_rows += 1
                continue

            exhibit, missing_fields = cls._build_exhibit_record(row_dict)
            exhibits.append(exhibit)
            if missing_fields:
                incomplete_rows.append({
                    "row_number": row_number,
                    "name": exhibit["name"],
                    "missing_fields": missing_fields,
                })

            if index == 1 or index == total_rows or index % 50 == 0:
                progress = 65 + int(index / total_rows * 30)
                report("processing_rows", progress, f"正在整理展品数据（第 {index}/{total_rows} 行）。")

        if not exhibits:
            raise HTTPException(status_code=400, detail="上传失败：未解析到有效展品数据。")

        return headers, exhibits, {
            "embedded_image_count": embedded_image_count,
            "skipped_blank_rows": skipped_blank_rows,
            "incomplete_rows": incomplete_rows,
            "incomplete_row_count": len(incomplete_rows),
        }

    @classmethod
    def _parse_saved_file(
        cls,
        file_path: Path,
        filename: str,
        task_id: str,
        user_id: str,
        uploaded_storage_paths: List[str],
        report: Callable[[str, int, str], None],
    ) -> Dict[str, Any]:
        lower_name = filename.lower()
        file_size = file_path.stat().st_size

        report("validating", 12, "正在检查模板和文件编码。")
        parse_meta = {
            "embedded_image_count": 0,
            "skipped_blank_rows": 0,
            "incomplete_rows": [],
            "incomplete_row_count": 0,
        }
        if lower_name.endswith(".csv"):
            report("reading_rows", 40, "CSV 模板校验通过，正在解析展品行数据。")
            headers, exhibits, parse_meta = cls._parse_csv(file_path)
            report("processing_rows", 92, "正在整理解析结果。")
        else:
            report("loading_excel", 18, "正在打开 Excel 文件。")
            headers, exhibits, parse_meta = cls._parse_xlsx(
                file_path,
                user_id,
                task_id,
                uploaded_storage_paths,
                report,
            )

        return {
            "headers": headers,
            "exhibits": exhibits,
            "total": len(exhibits),
            "file_name": filename,
            "meta": {
                "file_size": file_size,
                "imported_count": len(exhibits),
                **parse_meta,
            },
        }

    @classmethod
    async def parse_upload(cls, file: UploadFile, user_id: str) -> Dict[str, object]:
        """
        兼容旧同步接口。
        现在内部仍然走异步任务，前端可直接轮询任务状态。
        """
        return await cls.create_parse_task(file, user_id)


exhibit_import_service = ExhibitImportService()
